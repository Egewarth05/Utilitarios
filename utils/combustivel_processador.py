import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
import os
import pandas as pd
import sys
import csv
import re
from openpyxl import Workbook

def to_float(val):
    try:
        return float(val.replace(",", "."))
    except:
        return None

def processar_combustivel(caminho_csv, valor_gasolina, valor_diesel, caminho_saida):
    # — estilos —
    fill_gasolina = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    fill_diesel   = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    thin_border   = Border(left=Side(style='thin'), right=Side(style='thin'),
                           top=Side(style='thin'), bottom=Side(style='thin'))
    alinhamento   = Alignment(horizontal="center", vertical="center")

    # — converte preços —
    vg = float(str(valor_gasolina).replace(",", "."))
    vd = float(str(valor_diesel).replace(",", "."))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Relatório Combustível"

    # Preenche preços na 1ª linha
    ws["B1"] = f"Gasolina: R$ {vg:.2f}"
    ws["E1"] = f"Diesel:   R$ {vd:.2f}"

    # Agora segue sua legenda  
    ws["B2"].fill = fill_gasolina
    ws["C2"]       = "→ Gasolina"
    ws["E2"].fill = fill_diesel
    ws["F2"]       = "→ Diesel"

    # Cabeçalho na linha 6
    headers = ["Data", "Produto", "Nº da Nota", "Valor Total", "Quantidade", "TOTAL:"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=6, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.border = thin_border
        cell.alignment = alinhamento

    # Lê todo o CSV em memória
    with open(caminho_csv, encoding="latin1") as f:
        reader = list(csv.reader(f, delimiter=";"))

    linha_saida = 7
    ultima_data = ""
    ultima_nota = ""

    # Cada registro real ocupa um bloco de 4 linhas:
    # - linha i: dados gerais (data, número)
    # - linha i+3: detalhamento (produto, qtde, valor)
    for i in range(7, len(reader) - 3):
        try:
            linha_sup  = reader[i]
            linha_det  = reader[i + 3]

            # extrai data e número (mantém o último válido caso a célula fique vazia)
            data_raw = linha_sup[1].strip() if len(linha_sup) > 1 else ""
            nota_raw = linha_sup[2].strip() if len(linha_sup) > 2 else ""
            if data_raw: ultima_data = data_raw
            if nota_raw: ultima_nota = nota_raw

            # produto, qtde e valor na linha de detalhe
            produto    = linha_det[1].strip()
            produto_lower = produto.lower()
            # Filtra linhas indesejadas (Renault, placa, chassi, etc)
            if 'renault' in produto_lower or 'placa' in produto_lower or 'chassi' in produto_lower:
                continue

            quantidade = linha_det[3].strip()
            valor_txt  = linha_det[5].strip()

            qtde = to_float(quantidade)
            if qtde is None:
                continue

            # define preço e cor
            if "gasolina" in produto.lower():
                preco = vg
                cor   = fill_gasolina
            elif "diesel" in produto.lower():
                preco = vd
                cor   = fill_diesel
            else:
                continue

            total = round(qtde * preco, 2)

            # escreve no Excel
            dados = [ultima_data, produto, ultima_nota, valor_txt, qtde, total]
            for col_idx, val in enumerate(dados, start=1):
                cell = ws.cell(row=linha_saida, column=col_idx, value=val)
                cell.border = thin_border
                cell.alignment = alinhamento
                if col_idx == 6:  # coluna TOTAL
                    cell.fill = cor
                    cell.number_format = '#,##0.00'
                elif isinstance(val, float):
                    cell.number_format = '#,##0.00'

            linha_saida += 1
        except:
            # ignora quaisquer linhas malformadas
            continue

    # ajusta largura das colunas
    for coluna in ['A','B','C','D','E','F']:
        ws.column_dimensions[coluna].width = 15

    wb.save(caminho_saida)
